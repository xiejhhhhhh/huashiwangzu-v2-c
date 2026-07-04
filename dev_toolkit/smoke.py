"""
smoke_all — 一键全回归
后端集测(probe/call_capability) + 前端UI(Playwright) + 汇总红绿矩阵.
只访问活栈(33000/5173), 不重启服务.
断言规则: 只认内层 success, 拒绝 or status==200 兜底.
"""

import asyncio
import io
import json
import os
import re
import struct
import sys
import time
import zlib
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import httpx

try:
    from dev_toolkit.asset_lifecycle_tools import (
        CONFIRM_CLEAN_TEST_DATA,
        cleanup_test_data_pollution,
    )
except ModuleNotFoundError:
    from asset_lifecycle_tools import (
        CONFIRM_CLEAN_TEST_DATA,
        cleanup_test_data_pollution,
    )

# ── 配置 ──────────────────────────────────────────────────────────────

BACKEND_BASE = "http://127.0.0.1:33000"
FRONTEND_BASE = "http://localhost:5173"
REPO_ROOT = Path(__file__).resolve().parent.parent

ACCOUNTS = {
    "admin": {"username": "何焜华", "password": "123rgE123", "user_id": 4},
    "editor": {"username": "editor", "password": "admin123", "role": "editor"},
    "viewer": {"username": "viewer", "password": "admin123", "role": "viewer"},
}

TS = int(time.time() * 1000)
results: list[dict[str, Any]] = []
model_fallback_observations: list[dict[str, Any]] = []
_pending_deletions: list[int] = []  # 延后到所有测试结束后统一删除，避免异步 kb_pipeline 争抢
_TOKEN_CACHE: dict[str, str] = {}

# ── 辅助 ──────────────────────────────────────────────────────────────

async def _ensure_token(role: str = "admin", *, force_refresh: bool = False) -> str:
    if not force_refresh and role in _TOKEN_CACHE:
        return _TOKEN_CACHE[role]

    acct = ACCOUNTS.get(role, ACCOUNTS["admin"])
    async with httpx.AsyncClient(base_url=BACKEND_BASE, timeout=10, trust_env=False) as client:
        resp = await client.post("/api/login", json={
            "username": acct["username"],
            "password": acct["password"],
        })
        data = resp.json()
        token = data.get("data", data).get("access_token") or data.get("access_token")
        if not token:
            raise RuntimeError(f"Login failed {role}: status={resp.status_code}, data={data}")
        _TOKEN_CACHE[role] = token
        return token


def _json_or_raw(resp: httpx.Response) -> dict:
    try:
        return resp.json()
    except Exception:
        return {"raw": resp.text[:500]}


def _non_empty_error(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    try:
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    except TypeError:
        return str(value)


def _legacy_code_failure(value: Any) -> bool:
    if value in (None, "", 0, "0"):
        return False
    if isinstance(value, bool):
        return value is not False
    if isinstance(value, int | float):
        return value != 0
    if isinstance(value, str):
        try:
            return float(value.strip()) != 0
        except ValueError:
            return False
    return False


def _semantic_failure_reason(payload: Any, *, _depth: int = 0) -> str | None:
    if _depth > 8:
        return None
    if isinstance(payload, str):
        text = payload.strip()
        if not text or text[0] not in "{[":
            return None
        try:
            payload = json.loads(text)
        except json.JSONDecodeError:
            return None
    if not isinstance(payload, dict):
        return None
    if payload.get("success") is False:
        return _non_empty_error(payload.get("error")) or "success=false"
    error = _non_empty_error(payload.get("error"))
    if error:
        return error
    status = payload.get("status")
    if isinstance(status, str) and status.lower() in {"failed", "error"}:
        return f"status={status}"
    if "code" in payload and _legacy_code_failure(payload.get("code")):
        return (
            _non_empty_error(payload.get("message"))
            or _non_empty_error(payload.get("msg"))
            or f"code={payload.get('code')}"
        )
    for key in ("data", "result"):
        inner = payload.get(key)
        if isinstance(inner, (dict, str)):
            reason = _semantic_failure_reason(inner, _depth=_depth + 1)
            if reason:
                return reason
    return None


def _http_envelope_ok(response: dict) -> bool:
    status = response.get("status")
    if isinstance(status, int) and not 200 <= status < 300:
        return False
    return _semantic_failure_reason(response.get("data", response)) is None


async def _request_with_auth(
    method: str,
    path: str,
    *,
    role: str = "admin",
    timeout: int = 30,
    **kwargs: Any,
) -> dict:
    token = await _ensure_token(role)
    headers = dict(kwargs.pop("headers", {}) or {})
    headers["Authorization"] = f"Bearer {token}"
    async with httpx.AsyncClient(base_url=BACKEND_BASE, timeout=timeout, trust_env=False) as client:
        resp = await client.request(method, path, headers=headers, **kwargs)
        if resp.status_code == 401:
            token = await _ensure_token(role, force_refresh=True)
            headers["Authorization"] = f"Bearer {token}"
            resp = await client.request(method, path, headers=headers, **kwargs)
        return {"status": resp.status_code, "data": _json_or_raw(resp)}


async def _upload_file(filename: str, content: bytes, mime: str, folder_id: str = "0") -> dict:
    result = await _request_with_auth(
        "POST",
        "/api/files/upload",
        timeout=30,
        files={"file": (filename, content, mime)},
        data={"folder_id": folder_id},
    )
    return result["data"]

async def _delete_file(file_id: int) -> bool:
    r = await probe("POST", "/api/files/delete", {"id": file_id, "type": "file"})
    return _cap_ok(r)


async def _schedule_delete(file_id: int) -> None:
    """延后删除：记录 file_id，所有测试结束后统一清理。"""
    _pending_deletions.append(file_id)


async def _flush_pending_deletions() -> int:
    """执行所有延后删除，返回删除成功数。"""
    ok = 0
    for fid in _pending_deletions:
        try:
            if await _delete_file(fid):
                ok += 1
        except Exception as e:
            print(f"  [WARN] cleanup delete file_id={fid} failed: {e}")
    _pending_deletions.clear()
    return ok


async def _await_queue_settle(
    baseline_pending: int = 0,
    timeout: int = 30,
    interval: float = 0.25,
    quiet_samples: int = 2,
) -> dict:
    """Wait for the task queue pending count to settle back to baseline (or timeout)."""
    print(f"  等待异步队列静默... (pending baseline={baseline_pending})")
    deadline = time.monotonic() + timeout
    last_state = {}
    last_log = 0.0
    quiet_seen = 0
    while time.monotonic() < deadline:
        state = await _read_queue_state()
        pend = state.get("pending", 0)
        last_state = state
        if pend <= baseline_pending:
            quiet_seen += 1
            if quiet_seen >= quiet_samples:
                state["_settled"] = True
                print(f"  队列静默: pending={pend} (基线 {baseline_pending})")
                return state
        else:
            quiet_seen = 0
        elapsed = timeout - (deadline - time.monotonic())
        if elapsed - last_log >= 1 or last_log == 0.0:
            print(f"  等待中... pending={pend} (已等 {elapsed:.0f}s/{timeout}s)")
            last_log = elapsed
        await asyncio.sleep(min(interval, max(0.0, deadline - time.monotonic())))
    print(f"  超时: pending 未归零 (最后状态 pending={last_state.get('pending', '?')})")
    last_state["_settled"] = False
    last_state["_settle_timed_out"] = True
    return last_state


async def _read_queue_state() -> dict:
    r = await probe("GET", "/api/tasks/worker/status")
    if r.get("status") != 200 or not _http_envelope_ok(r):
        raise RuntimeError(f"Queue status probe failed: status={r.get('status')}, data={r.get('data')}")
    return r.get("data", {}).get("data", r.get("data", {}))


async def _quick_queue_quiet_probe(
    baseline_pending: int = 0,
    samples: int = 2,
    interval: float = 0.2,
) -> dict:
    last_state = {}
    for idx in range(samples):
        last_state = await _read_queue_state()
        if last_state.get("pending", 0) > baseline_pending:
            last_state["_quick_quiet"] = False
            return last_state
        if idx < samples - 1:
            await asyncio.sleep(interval)
    last_state["_quick_quiet"] = True
    return last_state

def _cap_ok(r: dict) -> bool:
    """Check capability inner success (data.data.success), fallback to data.success."""
    if not _http_envelope_ok(r):
        return False
    data = r.get("data", {})
    inner = data.get("data")
    if isinstance(inner, dict) and "success" in inner:
        return bool(inner.get("success"))
    return bool(data.get("success"))


def _cap_payload(r: dict) -> Any:
    """Unwrap /api/modules/call and module-level unified envelopes."""
    payload: Any = r.get("data", {})
    while isinstance(payload, dict) and "data" in payload:
        next_payload = payload.get("data")
        if next_payload is payload:
            break
        payload = next_payload
    return payload


def _relative_artifact_path(path_value: str) -> str:
    path = Path(path_value)
    if not path.is_absolute():
        return path_value
    try:
        return str(path.relative_to(REPO_ROOT))
    except ValueError:
        return str(path)


def _collect_playwright_failures(suite: dict[str, Any], parents: list[str] | None = None) -> list[dict[str, Any]]:
    parents = parents or []
    title = str(suite.get("title") or "").strip()
    next_parents = [*parents, title] if title else parents
    failures: list[dict[str, Any]] = []
    for spec in suite.get("specs", []) or []:
        if not isinstance(spec, dict):
            continue
        spec_title = str(spec.get("title") or "").strip()
        for test in spec.get("tests", []) or []:
            if not isinstance(test, dict):
                continue
            results_for_test = [r for r in (test.get("results") or []) if isinstance(r, dict)]
            last_result = results_for_test[-1] if results_for_test else {}
            status = str(test.get("status") or last_result.get("status") or "")
            if status not in {"failed", "timedOut", "interrupted", "unexpected"}:
                continue
            attachments = []
            for result_item in results_for_test:
                for attachment in result_item.get("attachments", []) or []:
                    if isinstance(attachment, dict) and attachment.get("path"):
                        attachments.append(_relative_artifact_path(str(attachment["path"])))
            error = last_result.get("error") if isinstance(last_result.get("error"), dict) else {}
            message = str(error.get("message") or last_result.get("error") or "")[:500]
            failures.append({
                "title": " › ".join([*next_parents, spec_title]).strip(" ›"),
                "status": status,
                "message": message,
                "artifacts": attachments[:6],
            })
    for child in suite.get("suites", []) or []:
        if isinstance(child, dict):
            failures.extend(_collect_playwright_failures(child, next_parents))
    return failures


def _parse_playwright_json(output: str, returncode: int, elapsed: float) -> dict[str, Any]:
    json_text = output.strip()
    if not json_text.startswith("{"):
        start = json_text.find("{")
        end = json_text.rfind("}")
        if start >= 0 and end > start:
            json_text = json_text[start:end + 1]
    try:
        payload = json.loads(json_text)
    except json.JSONDecodeError:
        return _parse_playwright_text(output, returncode, elapsed)

    stats = payload.get("stats") if isinstance(payload.get("stats"), dict) else {}
    passed = int(stats.get("expected") or 0)
    failed = int(stats.get("unexpected") or 0)
    skipped = int(stats.get("skipped") or 0)
    flaky = int(stats.get("flaky") or 0)
    failures: list[dict[str, Any]] = []
    for suite in payload.get("suites", []) or []:
        if isinstance(suite, dict):
            failures.extend(_collect_playwright_failures(suite))
    artifact_paths: list[str] = []
    for failure in failures:
        for artifact in failure.get("artifacts", []) or []:
            if artifact not in artifact_paths:
                artifact_paths.append(str(artifact))
    test_results_dir = REPO_ROOT / "frontend" / "test-results"
    if test_results_dir.exists():
        artifact_paths.append("frontend/test-results")
    status = "pass" if returncode == 0 and failed == 0 else "fail"
    return {
        "status": status,
        "returncode": returncode,
        "duration_seconds": round(elapsed, 3),
        "passed": passed,
        "failed": failed,
        "skipped": skipped,
        "flaky": flaky,
        "failed_tests": failures[:10],
        "artifact_paths": artifact_paths[:20],
        "reporter": "json",
    }


def _parse_playwright_text(output: str, returncode: int, elapsed: float) -> dict[str, Any]:
    passed_match = re.search(r"(\d+)\s*passed", output)
    failed_match = re.search(r"(\d+)\s*failed", output)
    skipped_match = re.search(r"(\d+)\s*(?:skipped|did not run)", output)
    fail_lines = [
        line.strip()
        for line in output.splitlines()
        if "FAIL" in line or "failed" in line.lower() or "✘" in line
    ]
    return {
        "status": "pass" if returncode == 0 else "fail",
        "returncode": returncode,
        "duration_seconds": round(elapsed, 3),
        "passed": int(passed_match.group(1)) if passed_match else 0,
        "failed": int(failed_match.group(1)) if failed_match else (1 if returncode != 0 else 0),
        "skipped": int(skipped_match.group(1)) if skipped_match else 0,
        "flaky": 0,
        "failed_tests": [{"title": line, "status": "failed", "message": ""} for line in fail_lines[-10:]],
        "artifact_paths": ["frontend/test-results"] if (REPO_ROOT / "frontend" / "test-results").exists() else [],
        "reporter": "text_fallback",
    }


def _classify_model_fallback_reason(warnings: list[str]) -> dict[str, Any]:
    text = "\n".join(warnings).lower()
    if any(token in text for token in ("401", "unauthorized", "forbidden", "api key", "apikey", "auth")):
        return {
            "category": "auth_config_debt",
            "retryable": False,
            "summary": "primary vision model auth/config failed; local fallback used",
        }
    if any(token in text for token in ("context", "too many tokens", "maximum", "request too large", "413")):
        return {
            "category": "context_too_large",
            "retryable": False,
            "summary": "vision payload exceeded context; compressed/local fallback used",
        }
    if "not configured" in text or "missing" in text:
        return {
            "category": "not_configured",
            "retryable": False,
            "summary": "vision model is not configured; local fallback used",
        }
    return {
        "category": "model_unavailable",
        "retryable": True,
        "summary": "vision model unavailable; local fallback used",
    }


def _record_model_fallback(source: str, payload: Any) -> dict[str, Any] | None:
    if not isinstance(payload, dict):
        return None
    strategy = payload.get("analysis_strategy") if isinstance(payload.get("analysis_strategy"), dict) else {}
    warnings = [str(item) for item in payload.get("warnings", []) if item]
    attempted = bool(strategy.get("vlm_attempted"))
    used = bool(strategy.get("vlm_used"))
    degraded = bool(strategy.get("degraded")) or (attempted and not used)
    if not attempted and not degraded:
        return None
    reason = _classify_model_fallback_reason(warnings)
    observation = {
        "source": source,
        "primary_model": str(strategy.get("primary_model") or "vision.primary"),
        "primary_failed": attempted and not used,
        "fallback_used": degraded,
        "fallback_model": "local_analysis",
        "final_success": bool(payload.get("description")),
        "failure_category": reason["category"] if degraded else "",
        "retryable": reason["retryable"],
        "summary": reason["summary"] if degraded else "vision model succeeded",
        "warnings": warnings[:5],
    }
    model_fallback_observations.append(observation)
    return observation


def _build_model_fallback_summary() -> dict[str, Any]:
    blockers = [item for item in model_fallback_observations if not item.get("final_success")]
    debts = [item for item in model_fallback_observations if item.get("fallback_used") and item.get("final_success")]
    if blockers:
        status = "BLOCKER"
    elif debts:
        status = "DEBT"
    else:
        status = "PASS"
    return {
        "status": status,
        "observations": model_fallback_observations,
        "fallback_used_count": len(debts),
        "blocker_count": len(blockers),
    }


def _load_backend_env(backend_dir: str) -> None:
    env_path = os.path.join(backend_dir, ".env")
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding="utf-8") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            os.environ.setdefault(key.strip(), value.strip().strip("\"'"))


async def _cleanup_scheduler_smoke_task(title: str) -> int:
    """Remove the scheduler task created by this smoke run."""
    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    backend_dir = os.path.join(repo_root, "backend")
    if backend_dir not in sys.path:
        sys.path.insert(0, backend_dir)
    _load_backend_env(backend_dir)

    from app.database import AsyncSessionLocal  # noqa: PLC0415
    from app.models.system import SystemTaskQueue  # noqa: PLC0415
    from sqlalchemy import delete  # noqa: PLC0415

    async with AsyncSessionLocal() as db:
        result = await db.execute(
            delete(SystemTaskQueue).where(
                SystemTaskQueue.module == "scheduler",
                SystemTaskQueue.task_type == "scheduled_agent_job",
                SystemTaskQueue.parameters.contains(title),
            )
        )
        await db.commit()
        return int(result.rowcount or 0)

def _cleanup_test_data_pollution() -> dict[str, Any]:
    return cleanup_test_data_pollution(
        REPO_ROOT,
        dry_run=False,
        limit=500,
        confirm=CONFIRM_CLEAN_TEST_DATA,
        reason=f"smoke_all:{TS}",
    )

def _make_png() -> bytes:
    """Minimal 2×2 red PNG."""
    width, height = 2, 2
    raw = b""
    for _ in range(height):
        raw += b"\x00"
        for _ in range(width):
            raw += b"\xff\x00\x00"
    def chunk(ctype: bytes, data: bytes) -> bytes:
        c = ctype + data
        return struct.pack(">I", len(data)) + c + struct.pack(">I", zlib.crc32(c) & 0xffffffff)
    ihdr = struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0)
    idat = zlib.compress(raw)
    return b"\x89PNG\r\n\x1a\n" + chunk(b"IHDR", ihdr) + chunk(b"IDAT", idat) + chunk(b"IEND", b"")

async def probe(method: str, path: str, body: dict | None = None, role: str = "admin") -> dict:
    return await _request_with_auth(method, path, role=role, timeout=30, json=body)

async def call_capability(module: str, action: str, params: dict | None = None, role: str = "admin") -> dict:
    body = {
        "target_module": module,
        "action": action,
        "parameters": params or {},
    }
    return await _request_with_auth("POST", "/api/modules/call", role=role, timeout=60, json=body)

def add_result(scenario: str, passed: bool, notes: str = "", status: str | None = None) -> None:
    result_status = status or ("PASS" if passed else "FAIL")
    results.append({"scenario": scenario, "passed": passed, "status": result_status, "notes": notes})
    marker = {"PASS": "G", "FAIL": "R", "SKIPPED": "S", "DEBT": "D"}.get(result_status, "?")
    print(f"  [{marker}] {scenario}: {notes[:120]}")


def _build_summary() -> dict[str, Any]:
    total = len(results)
    skipped = sum(1 for r in results if r.get("status") == "SKIPPED")
    debt = sum(1 for r in results if r.get("status") == "DEBT")
    failed = sum(1 for r in results if r.get("status") == "FAIL" or not r.get("passed"))
    passed = total - skipped - debt - failed
    verdict = "FAIL" if failed else ("PASS_WITH_DEBT" if skipped or debt else "PASS")
    ui_summary = next((r.get("data") for r in results if r.get("scenario") == "UI 集测 (Playwright)"), None)
    return {
        "verdict": verdict,
        "clean_pass": verdict == "PASS",
        "has_debt": verdict == "PASS_WITH_DEBT",
        "counts": {"total": total, "passed": passed, "failed": failed, "skipped": skipped, "debt": debt},
        "failed_scenarios": [r["scenario"] for r in results if r.get("status") == "FAIL" or not r.get("passed")],
        "debt_scenarios": [r["scenario"] for r in results if r.get("status") == "DEBT"],
        "skipped_scenarios": [r["scenario"] for r in results if r.get("status") == "SKIPPED"],
        "ui": ui_summary,
        "model_fallback": _build_model_fallback_summary(),
    }


def _new_failed_delta(failed_now: int, baseline_failed: int) -> int:
    """Only count failures added by this smoke run; external cleanup is not a failure."""
    return max(0, int(failed_now or 0) - int(baseline_failed or 0))


def _no_new_queue_failures(failed_now: int, baseline_failed: int) -> bool:
    return _new_failed_delta(failed_now, baseline_failed) == 0

# ── A. 框架主链路 ──────────────────────────────────────────────────

async def test_a():
    print("\n═══════════════════ A. 框架主链路 ═══════════════════\n")

    # A1 三角色登录
    for role in ("admin", "editor", "viewer"):
        try:
            t = await _ensure_token(role)
            add_result(f"A1 登录 {role}", bool(t), "token 签发成功")
        except Exception as e:
            add_result(f"A1 登录 {role}", False, str(e))

    # A3 desktop-apps
    r = await probe("GET", "/api/desktop/apps")
    data = r.get("data", {})
    apps = data.get("data", []) if isinstance(data, dict) else data
    if isinstance(apps, list):
        add_result("A3 desktop-apps 列出", len(apps) > 0, f"{len(apps)} 个应用")
    else:
        add_result("A3 desktop-apps 列出", False, f"非数组: {str(apps)[:100]}")

    # A4 上传/下载
    try:
        up = await _upload_file(f"smoke-{TS}.txt", b"hello smoke", "text/plain")
        if up.get("success"):
            file_id = up["data"]["id"]
            add_result("A4 上传文件", True, f"file_id={file_id}")
            down = await probe("GET", f"/api/files/download/{file_id}")
            add_result("A4 下载文件", down["status"] == 200, f"status={down['status']}")
            await _schedule_delete(file_id)
        else:
            add_result("A4 上传文件", False, up.get("error", "unknown"))
    except Exception as e:
        add_result("A4 上传/下载", False, str(e))

    # A5 recycle
    try:
        up = await _upload_file(
            f"recycle-{TS}.bin",
            b"to recycle",
            "application/octet-stream",
        )
        if up.get("success"):
            fid = up["data"]["id"]
            del_r = await probe("POST", "/api/files/delete", {"id": fid, "type": "file"})
            ok = _cap_ok(del_r)
            add_result("A5 删除到回收站", ok, f"file_id={fid}")
            # 回收站文件保留用于还原测试，不由 schedule_delete 清理

            list_r = await probe("GET", "/api/recycle/list?page=1&page_size=50")
            resp = list_r.get("data", {})
            items = resp.get("data", []) if isinstance(resp, dict) else resp
            found = any(i.get("origin_id") == fid for i in items)
            add_result("A5 回收站可见", found, f"in recycle={found}")

            recycle_id = None
            for i in items:
                if i.get("origin_id") == fid:
                    recycle_id = i.get("id")
                    break
            if recycle_id:
                rest = await probe("POST", "/api/recycle/restore", {"id": recycle_id, "item_type": "file"})
                ok = _cap_ok(rest)
                add_result("A5 还原", ok, "restore OK")
                if ok:
                    await _schedule_delete(fid)
        else:
            add_result("A5 recycle", False, "upload failed")
    except Exception as e:
        add_result("A5 recycle", False, str(e))

    # A7 dashboard stats
    r = await probe("GET", "/api/dashboard/stats")
    data = r.get("data", {}).get("data", {})
    has_stats = bool(data.get("total_files") is not None)
    add_result("A7 dashboard stats", has_stats, f"{str(data)[:100]}")

# ── B. 知识库 + 解析器 ────────────────────────────────────────────

async def test_b():
    print("\n═══════════════ B. 知识库 + 解析器 ═══════════════\n")

    r = await call_capability("knowledge", "search", {"query": "test", "top_k": 3})
    ok = _cap_ok(r)
    add_result("B1 knowledge search", ok, str(r.get("data", {}))[:120])

    # text-parser via upload + call
    try:
        up = await _upload_file(f"smoke-b2-{TS}.txt", b"Hello smoke test file content", "text/plain")
        if up.get("success"):
            fid = up["data"]["id"]
            r = await call_capability("text-parser", "parse", {"file_id": fid})
            ok = _cap_ok(r)
            add_result("B2 text-parser parse", ok, str(r.get("data", {}))[:120])
            await _schedule_delete(fid)
        else:
            add_result("B2 text-parser parse", False, f"upload failed: {up.get('error','?')}")
    except Exception as e:
        add_result("B2 text-parser parse", False, str(e))

    # pdf-parser via upload + call
    try:
        raw_pdf = b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 4 0 R/Resources<</Font<</F1 5 0 R>>>>>>endobj\n4 0 obj<</Length 44>>stream\nBT /F1 12 Tf 50 700 Td (hello) Tj ET\nendstream\nendobj\n5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj\nxref\n0 6\n0000000000 65535 f \n0000000009 00000 n \n0000000058 00000 n \n0000000115 00000 n \n0000000266 00000 n \n0000000360 00000 n \ntrailer<</Size 6/Root 1 0 R>>\nstartxref\n437\n%%EOF"
        up = await _upload_file(f"smoke-b2p-{TS}.pdf", raw_pdf, "application/pdf")
        if up.get("success"):
            fid = up["data"]["id"]
            r = await call_capability("pdf-parser", "parse", {"file_id": fid})
            ok = _cap_ok(r)
            add_result("B2 pdf-parser parse", ok, str(r.get("data", {}))[:120])
            await _schedule_delete(fid)
        else:
            add_result("B2 pdf-parser parse", False, f"upload failed: {up.get('error','?')}")
    except Exception as e:
        add_result("B2 pdf-parser parse", False, str(e))

    # image-vision describe: upload real PNG → get file_id → describe
    try:
        png = _make_png()
        up = await _upload_file(f"smoke-img-{TS}.png", png, "image/png")
        if up.get("success"):
            fid = up["data"]["id"]
            r = await call_capability("image-vision", "describe", {"file_id": fid})
            ok = _cap_ok(r)
            payload = _cap_payload(r)
            blocks = payload.get("blocks", []) if isinstance(payload, dict) else []
            _record_model_fallback("image-vision:auto", payload)
            add_result("B2 image-vision describe", ok, f"blocks={len(blocks)}")
            await _schedule_delete(fid)
        else:
            add_result("B2 image-vision describe", False, f"upload failed: {up.get('error','?')}")
    except Exception as e:
        add_result("B2 image-vision describe", False, str(e))

    # Semantic mode intentionally exercises the model/fallback lane. A model 401,
    # missing key, or context issue is debt only if local fallback still succeeds.
    try:
        png = _make_png()
        up = await _upload_file(f"smoke-vlm-{TS}.png", png, "image/png")
        if up.get("success"):
            fid = up["data"]["id"]
            r = await call_capability("image-vision", "describe", {
                "file_id": fid,
                "analysis_mode": "semantic",
                "prompt": "smoke gate semantic fallback probe",
            })
            ok = _cap_ok(r)
            payload = _cap_payload(r)
            observation = _record_model_fallback("image-vision:semantic", payload)
            if ok and observation and observation.get("fallback_used"):
                add_result(
                    "B3 image-vision model fallback",
                    True,
                    str(observation.get("summary") or "")[:160],
                    status="DEBT",
                )
            elif ok:
                add_result("B3 image-vision model fallback", True, "vision model succeeded without fallback")
            else:
                add_result("B3 image-vision model fallback", False, str(r.get("data", {}))[:160])
            await _schedule_delete(fid)
        else:
            add_result("B3 image-vision model fallback", False, f"upload failed: {up.get('error','?')}")
    except Exception as e:
        add_result("B3 image-vision model fallback", False, str(e))

# ── C. Agent 全链路 ───────────────────────────────────────────────

async def test_c():
    print("\n═══════════════════ C. Agent 全链路 ═══════════════════\n")

    r = await call_capability("memory", "overview_stats", {})
    ok = _cap_ok(r)
    add_result("C1 memory overview_stats", ok, str(r.get("data", {}))[:120])

    r = await call_capability("memory", "overview_stats", {})
    ok = _cap_ok(r)
    add_result("C2 memory overview_stats", ok, str(r.get("data", {}))[:120])

# ── D. 查看器 ─────────────────────────────────────────────────────

async def test_d():
    print("\n═══════════════════ D. 查看器 ═══════════════════\n")

    # D1 docs-open: POST with file_id (GET returns 405)
    try:
        up = await _upload_file(f"smoke-doc-{TS}.txt", b"hello docs open", "text/plain")
        if up.get("success"):
            fid = up["data"]["id"]
            r = await call_capability("docs-open", "open", {"file_id": fid})
            ok = _cap_ok(r)
            embed_url = r.get("data", {}).get("data", {}).get("embed_url", "")
            add_result("D1 docs-open", ok, f"has_embed={bool(embed_url)}")
            await _schedule_delete(fid)
        else:
            add_result("D1 docs-open", False, f"upload failed: {up.get('error','?')}")
    except Exception as e:
        add_result("D1 docs-open", False, str(e))

    # D2 excel-engine parse: upload real xlsx → get file_id → parse
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "col1"
        ws["B1"] = "col2"
        ws["A2"] = 1
        ws["B2"] = 2
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "data"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_data = buf.getvalue()
        up = await _upload_file(f"smoke-xl-{TS}.xlsx", xlsx_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if up.get("success"):
            fid = up["data"]["id"]
            r = await call_capability("excel-engine", "parse", {"file_id": fid})
            ok = _cap_ok(r)
            sheets = r.get("data", {}).get("data", {}).get("all_sheets", [])
            add_result("D2 excel-engine parse", ok and len(sheets) >= 2, f"sheets={sheets}")
            await _schedule_delete(fid)
        else:
            add_result("D2 excel-engine parse", False, f"upload failed: {up.get('error','?')}")
    except Exception as e:
        add_result("D2 excel-engine parse", False, str(e))

# ── E. 工具/生成类 ─────────────────────────────────────────────────

async def test_e():
    print("\n═══════════════════ E. 工具/生成类 ═══════════════════\n")

    r = await call_capability("image-gen", "list_templates", {})
    ok = _cap_ok(r)
    templates = r.get("data", {}).get("data", {}).get("templates", [])
    add_result("E1 image-gen list_templates", ok, f"{len(templates)} templates")

    # office-gen docx: filename (not file_name) + block content
    r = await call_capability("office-gen", "docx", {
        "filename": f"smoke-{TS}",
        "content": [{"type": "heading", "text": "标题"}, {"type": "paragraph", "text": "正文"}],
    })
    ok = _cap_ok(r)
    docx_id = r.get("data", {}).get("data", {}).get("file_id")
    add_result("E2 office-gen docx", ok, f"file_id={docx_id}")
    if docx_id:
        await _schedule_delete(docx_id)

    # office-gen xlsx
    r = await call_capability("office-gen", "xlsx", {
        "filename": f"smoke-{TS}",
        "sheets": [{"name": "Sheet1", "rows": [["a", "b"], ["1", "2"]]}],
    })
    ok = _cap_ok(r)
    xlsx_id = r.get("data", {}).get("data", {}).get("file_id")
    add_result("E3 office-gen xlsx", ok, f"file_id={xlsx_id}")
    if xlsx_id:
        await _schedule_delete(xlsx_id)

    # desktop-tools
    r = await call_capability("desktop-tools", "list_files", {"folder_id": 0})
    ok = _cap_ok(r)
    add_result("E5 desktop-tools list_files", ok, str(r.get("data", {}))[:120])

    # im send: need conversation_id, create if none exists
    try:
        r = await probe("GET", "/api/im/conversations")
        convs = r.get("data", {}).get("data", [])
        if not convs:
            r2 = await probe("POST", "/api/im/messages", {"target_user_id": 3, "content": f"smoke bootstrap {TS}"})
            convs2 = _cap_payload(r2)
            conv_id = convs2.get("conversation_id") if isinstance(convs2, dict) else None
        else:
            conv_id = convs[0]["id"]
        if conv_id:
            r = await call_capability("im", "send", {"conversation_id": conv_id, "content": f"smoke test {TS}"})
            ok = _cap_ok(r)
            msg_id = r.get("data", {}).get("data", {}).get("message_id")
            add_result("E7 im send", ok, f"message_id={msg_id}")
        else:
            add_result("E7 im send", False, "could not create conversation")
    except Exception as e:
        add_result("E7 im send", False, str(e))

    scheduler_title = f"smoke-{TS}"
    scheduled_at = (datetime.now(timezone.utc) + timedelta(days=365)).isoformat()
    r = await call_capability(
        "scheduler",
        "create",
        {
            "title": scheduler_title,
            "action_description": "smoke test scheduler noop",
            "scheduled_at": scheduled_at,
        },
    )
    ok = _cap_ok(r)
    task_payload = _cap_payload(r)
    task_id = task_payload.get("id") if isinstance(task_payload, dict) else None
    cleaned = 0
    if ok and task_id:
        cancel_result = await call_capability("scheduler", "cancel", {"task_id": task_id})
        ok = _cap_ok(cancel_result)
        cleaned = await _cleanup_scheduler_smoke_task(scheduler_title)
        add_result("E8 scheduler create/cancel", ok and cleaned == 1, f"task_id={task_id}, cleaned={cleaned}")
    else:
        add_result("E8 scheduler create/cancel", ok, str(r.get("data", {}))[:120])

# ── 前端 UI 集测 ──────────────────────────────────────────────────

async def test_ui():
    print("\n═══════════════════ 前端 UI 集测 ═══════════════════\n")
    frontend_dir = REPO_ROOT / "frontend"
    env = os.environ.copy()
    env["PLAYWRIGHT_EXTERNAL_SERVER"] = "1"

    try:
        proc = await asyncio.create_subprocess_exec(
            "npm", "run", "test:browser", "--", "--reporter=json", "--trace", "retain-on-failure",
            cwd=str(frontend_dir),
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            env=env,
        )
        started = time.monotonic()
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=240)
        elapsed = time.monotonic() - started
        output = stdout.decode(errors="replace") + stderr.decode(errors="replace")
        passed = proc.returncode == 0
        ui_summary = _parse_playwright_json(output, int(proc.returncode or 0), elapsed)
        add_result(
            "UI 集测 (Playwright)",
            passed,
            (
                f"exit={proc.returncode}, passed={ui_summary['passed']}, "
                f"failed={ui_summary['failed']}, skipped={ui_summary['skipped']}, "
                f"artifacts={len(ui_summary['artifact_paths'])}"
            ),
        )
        results[-1]["data"] = ui_summary
        if not passed:
            for failure in ui_summary.get("failed_tests", [])[:5]:
                print(f"    {failure.get('title', '?')}: {failure.get('message', '')[:160]}")
    except asyncio.TimeoutError:
        summary = {
            "status": "timeout",
            "returncode": None,
            "duration_seconds": 240,
            "passed": 0,
            "failed": 1,
            "skipped": 0,
            "flaky": 0,
            "failed_tests": [{"title": "Playwright run", "status": "timedOut", "message": "timeout (>240s)"}],
            "artifact_paths": ["frontend/test-results"] if (REPO_ROOT / "frontend" / "test-results").exists() else [],
            "reporter": "timeout",
        }
        add_result("UI 集测 (Playwright)", False, "超时(>240s)")
        results[-1]["data"] = summary
    except FileNotFoundError:
        summary = {
            "status": "unavailable",
            "returncode": None,
            "duration_seconds": 0,
            "passed": 0,
            "failed": 0,
            "skipped": 0,
            "flaky": 0,
            "failed_tests": [],
            "artifact_paths": [],
            "reporter": "unavailable",
            "reason": "npm not found",
        }
        add_result("UI 集测 (Playwright)", True, "npm not found; UI environment unavailable", status="DEBT")
        results[-1]["data"] = summary
    except Exception as e:
        summary = {
            "status": "error",
            "returncode": None,
            "duration_seconds": 0,
            "passed": 0,
            "failed": 1,
            "skipped": 0,
            "flaky": 0,
            "failed_tests": [{"title": "Playwright run", "status": "failed", "message": str(e)[:500]}],
            "artifact_paths": [],
            "reporter": "exception",
        }
        add_result("UI 集测 (Playwright)", False, str(e))
        results[-1]["data"] = summary

# ── 健康检查 ──────────────────────────────────────────────────────────

async def health_check():
    print("\n═══════════════════ 健康检查 ═══════════════════\n")
    try:
        r = await probe("GET", "/api/health")
        h = r.get("data", {}).get("data", r.get("data", {}))
        ok = _http_envelope_ok(r) and h.get("status") == "ok"
        module_errors = h.get("module_errors")
        add_result("后端 health", ok, f"db={h.get('database')}, module_errors={module_errors}")
    except Exception as e:
        add_result("后端 health", False, str(e))

    try:
        async with httpx.AsyncClient(timeout=5, trust_env=False) as cli:
            r = await cli.get("http://127.0.0.1:30000/health")
            ok = r.status_code == 200
            add_result("bge-m3 嵌入服务", ok, f"status={r.status_code}")
    except Exception:
        add_result("bge-m3 嵌入服务", False, "unreachable (some features may degrade)")

# ── 主函数 ────────────────────────────────────────────────────────────

async def main():
    model_fallback_observations.clear()
    print("smoke_all — 一键全回归")
    print(f"时间: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}")
    print(f"后端: {BACKEND_BASE}  前端: {FRONTEND_BASE}")

    # Capture the queue baseline before business steps create async work.
    init_state = await _read_queue_state()
    pre_failed = init_state.get("failed", 0)
    pre_pending = init_state.get("pending", 0)
    print(f"初始队列基线: failed={pre_failed}, pending={pre_pending}")

    for group_name, group in [
        ("健康检查", health_check),
        ("A. 框架主链路", test_a),
        ("B. 知识库 + 解析器", test_b),
        ("C. Agent 全链路", test_c),
        ("D. 查看器", test_d),
        ("E. 工具/生成类", test_e),
    ]:
        try:
            await group()
        except Exception as exc:
            add_result(group_name, False, f"group crashed: {exc}")
    if not os.environ.get("SMOKE_SKIP_UI"):
        try:
            await test_ui()
        except Exception as exc:
            add_result("UI 集测 (Playwright)", False, f"group crashed: {exc}")
    else:
        add_result("UI 集测 (Playwright)", True, "SMOKE_SKIP_UI=1，前端 UI 集测本轮跳过", status="SKIPPED")

    # ── 清理 & 异步队列验证 ──
    print("\n═══════════════════ 清理 + 异步队列验证 ═══════════════════\n")

    # Wait for tasks created by this smoke run to drain back to the pre-run baseline.
    after_business = await _await_queue_settle(baseline_pending=pre_pending, timeout=30)

    # 延后删除所有测试文件
    cleanup_count = len(_pending_deletions)
    deleted = await _flush_pending_deletions()
    print(f"  清理: 删除了 {deleted} 个测试文件")

    try:
        pollution_cleanup = _cleanup_test_data_pollution()
        selected_files = int(pollution_cleanup.get("selected_files") or 0)
        archived_docs = int(pollution_cleanup.get("archived_documents") or 0)
        archived_packages = int(pollution_cleanup.get("archived_packages") or 0)
        deleted_file_rows = int(pollution_cleanup.get("deleted_file_rows") or 0)
        physical_errors = pollution_cleanup.get("physical_delete_errors") or []
        cleanup_ok = bool(pollution_cleanup.get("success")) and not physical_errors
        add_result(
            "Z3 测试数据污染清理",
            cleanup_ok,
            (
                f"selected={selected_files}, deleted_file_rows={deleted_file_rows}, "
                f"archived_docs={archived_docs}, archived_packages={archived_packages}, "
                f"physical_errors={len(physical_errors)}"
            ),
        )
    except Exception as exc:
        add_result("Z3 测试数据污染清理", False, str(exc))

    # 清理后先短间隔探测；未产生异步积压时跳过第二轮长 settle。
    if cleanup_count == 0:
        print("  清理队列为空，跳过第二轮异步队列等待")
        final = after_business
    else:
        post_cleanup = await _quick_queue_quiet_probe(baseline_pending=pre_pending)
        if post_cleanup.get("_quick_quiet"):
            print("  清理后队列保持静默，跳过第二轮异步队列等待")
            final = post_cleanup
        else:
            final = await _await_queue_settle(baseline_pending=pre_pending, timeout=30)

    # 查最终异步队列状态
    failed_now = final.get("failed", 0)
    pending_now = final.get("pending", 0)
    oldest = final.get("oldest_waiting_seconds", 0) or 0
    settle_timed_out = bool(final.get("_settle_timed_out"))
    new_failures = _new_failed_delta(failed_now, pre_failed)
    add_result("Z1 异步队列无意外新增失败", _no_new_queue_failures(failed_now, pre_failed),
               f"failed: {pre_failed}(业务前基线) → {failed_now}(终), 新增={new_failures}, "
               f"清理文件数={cleanup_count}")
    add_result("Z2 异步队列积压可解释", pending_now <= 5 and not settle_timed_out,
               f"pending={pending_now}, oldest_waiting={oldest}s, settle_timeout={settle_timed_out}")
    print("\n" + "=" * 60)
    print("  红绿矩阵")
    print("=" * 60)
    summary = _build_summary()
    counts = summary["counts"]
    total = counts["total"]
    passed = counts["passed"]
    failed = counts["failed"]
    skipped = counts["skipped"]
    debt = counts["debt"]

    print(f"\n{'场景':<40} {'结果':>6} {'备注'}")
    print("-" * 80)
    for r in results:
        status = r.get("status", "PASS" if r["passed"] else "FAIL")
        marker = {"PASS": "G", "FAIL": "R", "SKIPPED": "S"}.get(status, "?")
        print(f"{r['scenario']:<40} {marker:>6}  {r['notes'][:100]}")

    print(f"\n总计: {total} 场景, G {passed} 通过, R {failed} 失败, D {debt} 债务, S {skipped} 跳过")
    print("SMOKE_JSON: " + json.dumps(summary, ensure_ascii=False))
    if failed > 0:
        print(f"{failed} 个场景失败, 请查看详情")
        sys.exit(1)
    if skipped > 0:
        print("通过但存在跳过项/债务，不是干净 PASS。")
    else:
        print("全绿!")

if __name__ == "__main__":
    asyncio.run(main())
