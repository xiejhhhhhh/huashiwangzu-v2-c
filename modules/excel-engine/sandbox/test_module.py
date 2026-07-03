"""Sandbox test for excel-engine module.

Tests the parsing, state management, and generation pipeline.
"""
import asyncio
import json
import os
import sys
import tempfile
from datetime import UTC, datetime

# Add module root to path (for `from backend.xxx` imports)
_MODULE_ROOT = os.path.dirname(os.path.dirname(__file__))  # modules/excel-engine/
_REPO_ROOT = os.path.abspath(os.path.join(_MODULE_ROOT, '..', '..'))
_BACKEND_ROOT = os.path.join(_REPO_ROOT, 'backend')
sys.path.insert(0, _MODULE_ROOT)
sys.path.insert(1, _BACKEND_ROOT)


def _load_backend_env() -> None:
    env_path = os.path.join(_BACKEND_ROOT, '.env')
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding='utf-8') as env_file:
        for raw_line in env_file:
            line = raw_line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            key, value = line.split('=', 1)
            os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


_load_backend_env()


def test_manifest_public_actions_match_registered_parameters():
    """Manifest discovery metadata must match the runtime capability registry."""
    # Importing the router registers excel-engine capabilities as it does in the app.
    import backend.router  # noqa: F401
    from app.services.module_registry import list_capabilities

    manifest_path = os.path.join(_MODULE_ROOT, "manifest.json")
    with open(manifest_path, encoding="utf-8") as manifest_file:
        manifest = json.load(manifest_file)

    manifest_actions = {
        item["action"]: {
            "min_role": item.get("min_role", "viewer"),
            "parameters": item.get("parameters") or {},
        }
        for item in manifest.get("public_actions", [])
    }
    registered_actions = {
        item["action"]: {
            "min_role": item.get("min_role", "viewer"),
            "parameters": item.get("parameters") or {},
        }
        for item in list_capabilities()
        if item.get("module") == "excel-engine"
    }

    assert manifest_actions == registered_actions
    print("  ✓ test_manifest_public_actions_match_registered_parameters passed")


# Test constants
TEST_DATA = {
    'cells': {
        'A1': '姓名', 'B1': '年龄', 'C1': '城市',
        'A2': '张三', 'B2': '28', 'C2': '北京',
        'A3': '李四', 'B3': '32', 'C3': '上海',
    },
    'styles': {
        'A1': {'bold': True, 'fontName': '微软雅黑', 'fontSize': 12, 'fillColor': '#4472C4', 'color': '#FFFFFF'},
    },
    'merges': {},
    'col_widths': {'A': 100, 'B': 60},
    'row_heights': {},
    'total_rows': 40,
    'total_cols': 10,
}


def test_address_tool():
    """Test address utility functions"""
    from backend.tool.address import col_letter, parse_address, rc_to_address

    assert parse_address('A1') == {'r': 1, 'c': 0}
    assert parse_address('B8') == {'r': 8, 'c': 1}
    assert parse_address('AA2') == {'r': 2, 'c': 26}

    assert rc_to_address(1, 0) == 'A1'
    assert rc_to_address(5, 1) == 'B5'
    assert rc_to_address(3, 26) == 'AA3'

    assert col_letter(0) == 'A'
    assert col_letter(25) == 'Z'
    assert col_letter(26) == 'AA'

    print('  ✓ test_address_tool passed')


def test_formula():
    """Test formula calculation"""
    from backend.tool.formula import calculate_formula

    cells = {'A1': '10', 'A2': '20', 'A3': '30', 'B1': '5'}
    assert calculate_formula('=SUM(A1:A3)', cells) == '60.0'
    assert calculate_formula('=AVERAGE(A1:A3)', cells) == '20.0'
    assert calculate_formula('=COUNT(A1:A3)', cells) == '3'
    assert calculate_formula('=MAX(A1:A3)', cells) == '30.0'
    assert calculate_formula('=MIN(A1:A3)', cells) == '10.0'
    assert calculate_formula('=A1+B1', cells) == '15'
    assert calculate_formula('=A1*B1', cells) == '50'
    assert calculate_formula('=A2/B1', cells) == '4.0'
    assert calculate_formula('hello', cells) == 'hello'

    print('  ✓ test_formula passed')


def test_state_manager():
    """Test state manager operations"""
    from backend.state.db_ops import empty_state
    from backend.state.manager import cell_get_style_ref, cell_get_text, cell_set_style_val, cell_set_text

    state = empty_state()
    assert state['total_rows'] == 40
    assert state['total_cols'] == 10

    # Cell operations
    cell_set_text(state, 'A1', 'Hello')
    assert cell_get_text(state, 'A1') == 'Hello'

    cell_set_text(state, 'A1', '')
    assert cell_get_text(state, 'A1') == ''

    # Style operations
    cell_set_style_val(state, 'B1', 'bold', True)
    assert cell_get_style_ref(state, 'B1')['bold'] is True

    print('  ✓ test_state_manager passed')


def test_xlsx_roundtrip():
    """Test XLSX generation and re-parsing"""
    from backend.engine.xlsx_generator import generate_xlsx

    fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)

    try:
        success = generate_xlsx(tmp_path, {'Sheet1': TEST_DATA})
        assert success, "XLSX generation failed"
        assert os.path.getsize(tmp_path) > 100, "Generated file too small"

        # Verify file is valid by re-opening with openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, read_only=True)
        ws = wb['Sheet1']
        assert ws is not None
        assert ws['A1'].value == '姓名'
        assert ws['A2'].value == '张三'
        wb.close()

        print('  ✓ test_xlsx_roundtrip passed')
    finally:
        os.unlink(tmp_path)


def test_csv_generation():
    """Test CSV generation"""
    from backend.engine.xlsx_generator import generate_csv

    fd, tmp_path = tempfile.mkstemp(suffix='.csv')
    os.close(fd)

    try:
        csv_content = generate_csv(tmp_path, TEST_DATA)
        assert '姓名' in csv_content
        assert '张三' in csv_content
        assert os.path.getsize(tmp_path) > 20

        print('  ✓ test_csv_generation passed')
    finally:
        os.unlink(tmp_path)


async def _async_test_edit_operations():
    """Test edit operations"""
    from backend.state.manager import cell_get_text
    from backend.table.edit import EditOperations

    state = dict(TEST_DATA)
    result = await EditOperations._input(state, 'test_key', ['A1'], {'value': '测试值'})
    assert result['code'] == 0
    assert state['cells']['A1'] == '测试值'

    result = await EditOperations._clear(state, 'test_key', ['A1', 'B1'], {'type': 'all'})
    assert result['code'] == 0
    assert cell_get_text(state, 'A1') == ''
    assert 'A1' not in state['cells']

    print('  ✓ test_edit_operations passed')


def test_edit_operations():
    asyncio.run(_async_test_edit_operations())


async def _async_test_style_operations():
    """Test style operations"""
    from backend.table.style_ops import StyleOperations

    state = dict(TEST_DATA)
    result = await StyleOperations._toggle_style(state, ['A1'], 'bold')
    assert result['code'] == 0
    # Toggle again should be off
    await StyleOperations._toggle_style(state, ['A1'], 'bold')
    assert state['styles'].get('A1', {}).get('bold') is False

    print('  ✓ test_style_operations passed')


def test_style_operations():
    asyncio.run(_async_test_style_operations())


async def _async_test_row_col_operations():
    """Test row/column operations"""
    from backend.table.row_col import RowColOperations

    state = dict(TEST_DATA)
    old_a2_value = state.get('cells', {}).get('A2', '')
    result = await RowColOperations._delete_row(state, ['A2'])
    assert result['code'] == 0
    # Row 2 deleted: old A2 value gone, cells above moved down
    assert state.get('cells', {}).get('A2', '') != old_a2_value
    assert state.get('total_rows', 0) == 39

    state = {'cells': {'A1': 'left', 'B1': 'right'}, 'styles': {}, 'merges': {}, 'total_rows': 40, 'total_cols': 10}
    result = await RowColOperations._insert_shift_right(state, ['A1'])
    assert result['code'] == 0
    assert state['cells'].get('B1') == 'left'
    assert state['cells'].get('C1') == 'right'

    result = await RowColOperations._delete_shift_right(state, ['B1'])
    assert result['code'] == 0
    assert state['cells'].get('B1') == 'right'

    state = {'cells': {'A1': 'top', 'A2': 'bottom'}, 'styles': {}, 'merges': {}, 'total_rows': 40, 'total_cols': 10}
    result = await RowColOperations._insert_shift_down(state, ['A1'])
    assert result['code'] == 0
    assert state['cells'].get('A2') == 'top'
    assert state['cells'].get('A3') == 'bottom'

    result = await RowColOperations._delete_shift_up(state, ['A2'])
    assert result['code'] == 0
    assert state['cells'].get('A2') == 'bottom'

    print('  ✓ test_row_col_operations passed')


def test_row_col_operations():
    asyncio.run(_async_test_row_col_operations())


async def _async_test_clipboard():
    """Test clipboard operations"""
    from backend.table.clipboard import ClipboardOperations

    state = dict(TEST_DATA)
    state['_clipboard'] = {}
    state['_clipboard_range'] = []

    result = ClipboardOperations._copy(state, ['A1', 'B1'])
    assert result['code'] == 0
    assert 'A1' in state.get('_clipboard', {})
    assert 'B1' in state.get('_clipboard', {})

    print('  ✓ test_clipboard passed')


def test_clipboard():
    asyncio.run(_async_test_clipboard())


async def _cleanup_db_state(db, state_prefix: str, file_name_prefix: str) -> None:
    from sqlalchemy import text

    wb_rows = await db.execute(
        text("SELECT id FROM excel_workbooks WHERE state_key LIKE :prefix"),
        {"prefix": f"{state_prefix}%"},
    )
    workbook_ids = [row[0] for row in wb_rows.fetchall()]
    if workbook_ids:
        sheet_rows = await db.execute(
            text("SELECT id FROM excel_sheets WHERE workbook_id = ANY(:workbook_ids)"),
            {"workbook_ids": workbook_ids},
        )
        sheet_ids = [row[0] for row in sheet_rows.fetchall()]
        if sheet_ids:
            await db.execute(text("DELETE FROM excel_cells WHERE sheet_id = ANY(:sheet_ids)"), {"sheet_ids": sheet_ids})
            await db.execute(text("DELETE FROM excel_col_widths WHERE sheet_id = ANY(:sheet_ids)"), {"sheet_ids": sheet_ids})
            await db.execute(text("DELETE FROM excel_row_heights WHERE sheet_id = ANY(:sheet_ids)"), {"sheet_ids": sheet_ids})
            await db.execute(text("DELETE FROM excel_history WHERE sheet_id = ANY(:sheet_ids)"), {"sheet_ids": sheet_ids})
            await db.execute(text("DELETE FROM excel_redo_stack WHERE sheet_id = ANY(:sheet_ids)"), {"sheet_ids": sheet_ids})
        await db.execute(text("DELETE FROM excel_sheets WHERE workbook_id = ANY(:workbook_ids)"), {"workbook_ids": workbook_ids})
        await db.execute(text("DELETE FROM excel_workbooks WHERE id = ANY(:workbook_ids)"), {"workbook_ids": workbook_ids})

    file_rows = await db.execute(
        text("SELECT id FROM framework_file_items WHERE name LIKE :prefix"),
        {"prefix": f"{file_name_prefix}%"},
    )
    file_ids = [row[0] for row in file_rows.fetchall()]
    if file_ids:
        await db.execute(text("DELETE FROM excel_versions WHERE file_id = ANY(:file_ids)"), {"file_ids": file_ids})
        await db.execute(text("DELETE FROM framework_file_items WHERE id = ANY(:file_ids)"), {"file_ids": file_ids})
    await db.commit()


async def _cleanup_db_workbooks_by_state_keys(db, state_keys: list[str]) -> None:
    from sqlalchemy import text

    if not state_keys:
        return
    wb_rows = await db.execute(
        text("SELECT id FROM excel_workbooks WHERE state_key = ANY(:state_keys)"),
        {"state_keys": state_keys},
    )
    workbook_ids = [row[0] for row in wb_rows.fetchall()]
    if not workbook_ids:
        return
    sheet_rows = await db.execute(
        text("SELECT id FROM excel_sheets WHERE workbook_id = ANY(:workbook_ids)"),
        {"workbook_ids": workbook_ids},
    )
    sheet_ids = [row[0] for row in sheet_rows.fetchall()]
    if sheet_ids:
        await db.execute(text("DELETE FROM excel_cells WHERE sheet_id = ANY(:sheet_ids)"), {"sheet_ids": sheet_ids})
        await db.execute(text("DELETE FROM excel_col_widths WHERE sheet_id = ANY(:sheet_ids)"), {"sheet_ids": sheet_ids})
        await db.execute(text("DELETE FROM excel_row_heights WHERE sheet_id = ANY(:sheet_ids)"), {"sheet_ids": sheet_ids})
        await db.execute(text("DELETE FROM excel_history WHERE sheet_id = ANY(:sheet_ids)"), {"sheet_ids": sheet_ids})
        await db.execute(text("DELETE FROM excel_redo_stack WHERE sheet_id = ANY(:sheet_ids)"), {"sheet_ids": sheet_ids})
    await db.execute(text("DELETE FROM excel_sheets WHERE workbook_id = ANY(:workbook_ids)"), {"workbook_ids": workbook_ids})
    await db.execute(text("DELETE FROM excel_workbooks WHERE id = ANY(:workbook_ids)"), {"workbook_ids": workbook_ids})
    await db.commit()


async def _get_two_user_ids(db) -> tuple[int, int]:
    from sqlalchemy import text

    result = await db.execute(text("SELECT id FROM framework_user_accounts ORDER BY id LIMIT 2"))
    user_ids = [int(row[0]) for row in result.fetchall()]
    if len(user_ids) < 2:
        raise AssertionError("At least two users are required for excel-engine isolation tests")
    return user_ids[0], user_ids[1]


async def _async_test_db_state_key_owner_isolation():
    from app.database import AsyncSessionLocal
    from backend.init_db import run_init
    from backend.state.db_ops import (
        find_or_create_sheet,
        find_or_create_workbook,
        find_workbook,
        read_state_full,
        sync_cells,
    )

    stamp = datetime.now(UTC).strftime("%Y%m%d%H%M%S%f")
    state_key = f"w3_state_isolation_{stamp}"
    async with AsyncSessionLocal() as db:
        await run_init(db)
        owner_a, owner_b = await _get_two_user_ids(db)
        await _cleanup_db_state(db, state_key, f"w3_excel_file_{stamp}")
        try:
            wb_a = await find_or_create_workbook(db, state_key, owner_id=owner_a)
            wb_b = await find_or_create_workbook(db, state_key, owner_id=owner_b)
            assert wb_a["id"] != wb_b["id"]
            assert (await find_workbook(db, state_key, owner_id=owner_a))["id"] == wb_a["id"]
            assert (await find_workbook(db, state_key, owner_id=owner_b))["id"] == wb_b["id"]

            sheet_a = await find_or_create_sheet(db, wb_a["id"], "Sheet1")
            await sync_cells(db, sheet_a["id"], {"A1": "owner-a"}, {}, {})

            state_a = await read_state_full(db, state_key, owner_id=owner_a)
            state_b = await read_state_full(db, state_key, owner_id=owner_b)
            assert state_a["cells"].get("A1") == "owner-a"
            assert state_b["cells"].get("A1") != "owner-a"
        finally:
            await _cleanup_db_state(db, state_key, f"w3_excel_file_{stamp}")

    print('  ✓ test_db_state_key_owner_isolation passed')


def test_db_state_key_owner_isolation():
    asyncio.run(_async_test_db_state_key_owner_isolation())


async def _async_test_version_restore_rejects_cross_file_version_id():
    from app.database import AsyncSessionLocal
    from backend.init_db import run_init
    from backend.router import _restore_version_capability
    from backend.state.db_ops import find_or_create_sheet, find_or_create_workbook
    from sqlalchemy import text

    stamp = datetime.now(UTC).strftime("%Y%m%d%H%M%S%f")
    file_name_prefix = f"w3_excel_file_{stamp}"
    created_file_ids: list[int] = []
    async with AsyncSessionLocal() as db:
        await run_init(db)
        owner_a, owner_b = await _get_two_user_ids(db)
        await _cleanup_db_state(db, f"w3_unused_{stamp}", file_name_prefix)
        try:
            file_a = await db.execute(
                text(
                    "INSERT INTO framework_file_items "
                    "(name, extension, size, owner_id, storage_path, mime_type, ref_count, deleted) "
                    "VALUES (:name, 'xlsx', 0, :owner_id, '', '', 1, false) RETURNING id"
                ),
                {"name": f"{file_name_prefix}_a", "owner_id": owner_a},
            )
            file_b = await db.execute(
                text(
                    "INSERT INTO framework_file_items "
                    "(name, extension, size, owner_id, storage_path, mime_type, ref_count, deleted) "
                    "VALUES (:name, 'xlsx', 0, :owner_id, '', '', 1, false) RETURNING id"
                ),
                {"name": f"{file_name_prefix}_b", "owner_id": owner_b},
            )
            file_id_a = int(file_a.scalar_one())
            file_id_b = int(file_b.scalar_one())
            created_file_ids = [file_id_a, file_id_b]

            wb_a = await find_or_create_workbook(db, f"knowledge_{file_id_a}", owner_id=owner_a)
            wb_b = await find_or_create_workbook(db, f"knowledge_{file_id_b}", owner_id=owner_b)
            await find_or_create_sheet(db, wb_a["id"], "Sheet1")
            await find_or_create_sheet(db, wb_b["id"], "Sheet1")

            version_a = await db.execute(
                text(
                    "INSERT INTO excel_versions "
                    "(file_id, version_name, file_size, snapshot_json, operation_steps, creator_id, created_at) "
                    "VALUES (:file_id, 'owner-a', 0, :snapshot, 1, :creator_id, NOW()) RETURNING id"
                ),
                {"file_id": file_id_a, "snapshot": json.dumps({"cells": {"A1": "a"}}), "creator_id": owner_a},
            )
            version_b = await db.execute(
                text(
                    "INSERT INTO excel_versions "
                    "(file_id, version_name, file_size, snapshot_json, operation_steps, creator_id, created_at) "
                    "VALUES (:file_id, 'owner-b', 0, :snapshot, 1, :creator_id, NOW()) RETURNING id"
                ),
                {"file_id": file_id_b, "snapshot": json.dumps({"cells": {"A1": "b"}}), "creator_id": owner_b},
            )
            version_id_a = int(version_a.scalar_one())
            version_id_b = int(version_b.scalar_one())
            await db.commit()

            restored = await _restore_version_capability(
                {"state_key": f"knowledge_{file_id_a}", "version_id": version_id_a},
                f"user:{owner_a}",
            )
            assert restored["restored"] is True

            try:
                await _restore_version_capability(
                    {"state_key": f"knowledge_{file_id_a}", "version_id": version_id_b},
                    f"user:{owner_a}",
                )
            except ValueError as exc:
                assert "Version not found" in str(exc)
            else:
                raise AssertionError("Cross-file version_id restore should fail")
        finally:
            await db.rollback()
            if created_file_ids:
                await _cleanup_db_workbooks_by_state_keys(
                    db,
                    [f"knowledge_{file_id}" for file_id in created_file_ids],
                )
                await _cleanup_db_state(db, f"w3_unused_{stamp}", file_name_prefix)

    print('  ✓ test_version_restore_rejects_cross_file_version_id passed')


def test_version_restore_rejects_cross_file_version_id():
    asyncio.run(_async_test_version_restore_rejects_cross_file_version_id())


async def _async_test_live_capability_history_versions_compile():
    from pathlib import Path

    import openpyxl
    from app.config import get_settings
    from app.database import AsyncSessionLocal
    from backend.init_db import run_init
    from backend.router import (
        _append_rows_capability,
        _compile_xlsx_capability,
        _dispatch,
        _import_file_to_workbook_capability,
        _list_versions_capability,
        _redo_capability,
        _restore_version_capability,
        _undo_capability,
        _update_range_capability,
    )
    from backend.state.db_ops import read_state_full
    from sqlalchemy import text

    stamp = datetime.now(UTC).strftime("%Y%m%d%H%M%S%f")
    file_name_prefix = f"w3_excel_file_{stamp}"
    upload_root = Path(get_settings().UPLOAD_DIR).resolve()
    storage_dir = Path("excel_engine_test")
    storage_path = storage_dir / f"{file_name_prefix}.xlsx"
    full_path = upload_root / storage_path
    full_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "seed"
    ws.column_dimensions["A"].width = 22
    ws.row_dimensions[1].height = 28
    wb.save(full_path)
    wb.close()

    created_file_ids: list[int] = []
    compiled_path = ""
    async with AsyncSessionLocal() as db:
        await run_init(db)
        owner_id, _ = await _get_two_user_ids(db)
        await _cleanup_db_state(db, f"unused_{stamp}", file_name_prefix)
        try:
            file_result = await db.execute(
                text(
                    "INSERT INTO framework_file_items "
                    "(name, extension, size, owner_id, storage_path, mime_type, ref_count, deleted) "
                    "VALUES (:name, 'xlsx', :size, :owner_id, :storage_path, "
                    "'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 1, false) "
                    "RETURNING id"
                ),
                {
                    "name": f"{file_name_prefix}_import",
                    "size": full_path.stat().st_size,
                    "owner_id": owner_id,
                    "storage_path": str(storage_path),
                },
            )
            file_id = int(file_result.scalar_one())
            created_file_ids.append(file_id)
            await db.commit()
            state_key = f"knowledge_{file_id}"
            caller = f"user:{owner_id}"

            imported = await _import_file_to_workbook_capability({"file_id": file_id}, caller)
            assert imported["state_key"] == state_key

            width_count = await db.execute(
                text(
                    "SELECT COUNT(*) FROM excel_col_widths cw "
                    "JOIN excel_sheets s ON s.id = cw.sheet_id "
                    "JOIN excel_workbooks w ON w.id = s.workbook_id "
                    "WHERE w.state_key = :state_key"
                ),
                {"state_key": state_key},
            )
            height_count = await db.execute(
                text(
                    "SELECT COUNT(*) FROM excel_row_heights rh "
                    "JOIN excel_sheets s ON s.id = rh.sheet_id "
                    "JOIN excel_workbooks w ON w.id = s.workbook_id "
                    "WHERE w.state_key = :state_key"
                ),
                {"state_key": state_key},
            )
            assert int(width_count.scalar_one()) > 0
            assert int(height_count.scalar_one()) > 0

            updated = await _update_range_capability(
                {"state_key": state_key, "start_row": 0, "start_col": 0, "rows": [["updated", "B"]]},
                caller,
            )
            assert updated["rows_updated"] == 1
            appended = await _append_rows_capability({"state_key": state_key, "rows": [["append-a", "append-b"]]}, caller)
            assert appended["start_row"] == 2

            state = await read_state_full(db, state_key, owner_id=owner_id)
            assert state["cells"].get("A1") == "updated"
            assert state["cells"].get("A2") == "append-a"
            assert state["cells"].get("B2") == "append-b"

            undo = await _undo_capability({"state_key": state_key}, caller)
            assert undo["success"] is True
            assert "A2" not in undo["cells"]
            redo_stack_count = await db.execute(
                text(
                    "SELECT COUNT(*) FROM excel_redo_stack rs "
                    "JOIN excel_sheets s ON s.id = rs.sheet_id "
                    "JOIN excel_workbooks w ON w.id = s.workbook_id "
                    "WHERE w.state_key = :state_key"
                ),
                {"state_key": state_key},
            )
            assert int(redo_stack_count.scalar_one()) == 1

            redo = await _redo_capability({"state_key": state_key}, caller)
            assert redo["success"] is True
            assert redo["cells"].get("A2") == "append-a"

            saved = await _dispatch(
                "export",
                "save_version",
                {"version_name": "manual-test"},
                state_key,
                "Sheet1",
                db,
                owner_id=owner_id,
            )
            assert saved["code"] == 0
            version_id = int(saved["version"]["id"])
            versions = await _list_versions_capability({"state_key": state_key}, caller)
            assert any(int(v["id"]) == version_id for v in versions["versions"])

            await _update_range_capability({"state_key": state_key, "rows": [["changed-after-version"]]}, caller)
            restored = await _restore_version_capability({"state_key": state_key, "version_id": version_id}, caller)
            assert restored["restored"] is True
            assert restored["cells"].get("A1") == "updated"
            assert restored["cells"].get("A2") == "append-a"

            compiled = await _compile_xlsx_capability({"state_key": state_key}, caller)
            compiled_path = compiled["file_path"]
            assert os.path.exists(compiled_path)
            assert os.path.getsize(compiled_path) > 100
        finally:
            await db.rollback()
            if compiled_path and os.path.exists(compiled_path):
                os.unlink(compiled_path)
            if created_file_ids:
                await _cleanup_db_workbooks_by_state_keys(db, [f"knowledge_{file_id}" for file_id in created_file_ids])
                await _cleanup_db_state(db, f"unused_{stamp}", file_name_prefix)
            if full_path.exists():
                full_path.unlink()

    print('  ✓ test_live_capability_history_versions_compile passed')


def test_live_capability_history_versions_compile():
    asyncio.run(_async_test_live_capability_history_versions_compile())


if __name__ == '__main__':
    print('\n=== Excel Engine Module Tests ===\n')
    test_manifest_public_actions_match_registered_parameters()
    test_address_tool()
    test_formula()
    test_state_manager()
    test_xlsx_roundtrip()
    test_csv_generation()
    test_edit_operations()
    test_style_operations()
    test_row_col_operations()
    test_clipboard()
    test_db_state_key_owner_isolation()
    test_version_restore_rejects_cross_file_version_id()
    test_live_capability_history_versions_compile()
    print('\n✓ All tests passed!\n')
