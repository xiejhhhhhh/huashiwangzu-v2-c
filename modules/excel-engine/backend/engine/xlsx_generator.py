"""XLSX generator - 1:1 from old 引擎/xlsx生成.php + 写入_*.php

Generates XLSX files from state data using openpyxl.
"""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Any


def _decode_ascii_escape(val: str) -> str:
    def _replace(m):
        return chr(int(m.group(1), 16))
    return re.sub(r'\[ASCII:0x([0-9A-F]{4})\]', _replace, val)


def _map_align(align: str) -> str | None:
    m = {'左': 'left', '居中': 'center', '右': 'right'}
    return m.get(align)


def _apply_style_to_cell(cell, style_info: dict[str, Any]):
    """Apply style dict to openpyxl cell - 1:1 from old 写入_样式.php"""
    kwargs: dict[str, Any] = {}

    if style_info.get('bold'):
        kwargs['bold'] = True
    if style_info.get('italic'):
        kwargs['italic'] = True
    if style_info.get('underline'):
        kwargs['underline'] = 'single'
    if style_info.get('strikethrough'):
        kwargs['strike'] = True
    if style_info.get('fontSize'):
        kwargs['size'] = style_info['fontSize']
    if style_info.get('fontName'):
        kwargs['name'] = style_info['fontName']
    if style_info.get('color'):
        c = style_info['color']
        if c.startswith('#'):
            kwargs['color'] = c[1:]
    if kwargs:
        cell.font = Font(**kwargs)

    if style_info.get('fillColor'):
        fc = style_info['fillColor']
        if fc.startswith('#'):
            cell.fill = PatternFill(start_color=fc[1:], end_color=fc[1:], fill_type='solid')

    if style_info.get('align'):
        ha = _map_align(style_info['align'])
        if ha:
            cell.alignment = Alignment(horizontal=ha, wrap_text=style_info.get('wrapText', False))
    elif style_info.get('wrapText'):
        cell.alignment = Alignment(wrap_text=True)

    if style_info.get('borderType'):
        bt = style_info['borderType']
        sides = {}
        if bt == 'all' or bt == 'outside':
            for s in ['left', 'right', 'top', 'bottom']:
                sides[s] = Side(style='thin', color='000000')
        cell.border = Border(**sides)


def generate_xlsx(output_path: str, all_sheet_data: dict[str, Any], filename: str = '') -> bool:
    """Generate XLSX file from state data - 1:1 from old 生成XLSX文件"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, data in all_sheet_data.items():
        ws = wb.create_sheet(title=sheet_name)

        cells = data.get('cells', {})
        styles = data.get('styles', {})
        merges = data.get('merges', {})
        col_widths = data.get('col_widths', {})
        row_heights = data.get('row_heights', {})
        total_rows = data.get('total_rows', 40)
        total_cols = data.get('total_cols', 10)

        # Write cells
        for addr, val in cells.items():
            m = re.match(r'([A-Z]+)(\d+)', addr)
            if not m:
                continue
            col_str = m.group(1)
            row_num = int(m.group(2))
            col_idx = 0
            for ch in col_str:
                col_idx = col_idx * 26 + (ord(ch) - 64)
            cell = ws.cell(row=row_num, column=col_idx)

            # Decode ASCII escapes
            if isinstance(val, str) and '[ASCII:' in val:
                val = _decode_ascii_escape(val)

            # Check if it's a formula
            if isinstance(val, str) and val.startswith('='):
                cell.value = val
            else:
                cell.value = val

            # Apply style
            if addr in styles:
                _apply_style_to_cell(cell, styles[addr])

        # Apply merge cells
        for merge_ref, merge_info in merges.items():
            ws.merge_cells(merge_ref)

        # Column widths
        for col_letter, width in col_widths.items():
            try:
                ws.column_dimensions[col_letter].width = width / 7
            except KeyError:
                pass

        # Row heights
        for row_str, height in row_heights.items():
            try:
                ws.row_dimensions[int(row_str)].height = height / 4
            except (ValueError, KeyError):
                pass

    wb.save(output_path)
    return True


def generate_csv(output_path: str, data: dict[str, Any], sheet_name: str = 'Sheet1') -> str:
    """Generate CSV content from state data"""
    import csv
    import io

    cells = data.get('cells', {})
    total_rows = data.get('total_rows', 40)
    total_cols = data.get('total_cols', 10)

    output = io.StringIO()
    writer = csv.writer(output)

    for r in range(1, total_rows + 1):
        row_data = []
        for c in range(total_cols):
            addr = ''
            t = c + 1
            while t > 0:
                t -= 1
                addr = chr(65 + (t % 26)) + addr
                t //= 26
            addr = f'{addr}{r}'
            val = cells.get(addr, '')
            row_data.append(val)
        writer.writerow(row_data)

    csv_content = output.getvalue()
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        f.write(csv_content)
    return csv_content
