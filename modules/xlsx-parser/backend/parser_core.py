"""Pure XLSX/CSV parsing logic for xlsx-parser."""
from __future__ import annotations

import csv
import io
from collections.abc import Iterable
from datetime import date, datetime, time
from itertools import zip_longest
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

MAX_ROWS_PER_SHEET = 5000
SUPPORTED_EXTS = {"xlsx", "csv"}
SCHEMA_VERSION = "1.0"
SOURCE_MODULE = "xlsx-parser"
PARSER_NAME = "xlsx-parser:parse"


class SpreadsheetParseError(ValueError):
    """Raised when the input is not a parseable supported spreadsheet."""


def _decode_text_bytes(raw: bytes) -> str:
    for encoding in ("utf-8", "utf-8-sig", "gbk", "gb2312", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _format_cell(value: object, formula_value: object = None) -> str:
    if value is None:
        if isinstance(formula_value, str) and formula_value.startswith("="):
            return formula_value.strip()
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date | time):
        return value.isoformat()
    return str(value).strip()


def _row_to_text(values: Iterable[object], formulas: Iterable[object] = ()) -> str:
    cells = [
        _format_cell(value, formula)
        for value, formula in zip_longest(values, formulas, fillvalue=None)
    ]
    return " | ".join(cells).rstrip()


def _source(file_id: int, ext: str) -> dict[str, object]:
    mime_type = "text/csv" if ext == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return {
        "format": ext,
        "module": SOURCE_MODULE,
        "file_id": file_id,
        "filename": None,
        "mime_type": mime_type,
    }


def _block(
    block_type: str,
    text: str,
    source_ref: dict[str, object],
    page: int | None = None,
    metadata: dict[str, object] | None = None,
) -> dict[str, object]:
    data: dict[str, object] = {"source_ref": source_ref}
    if metadata:
        data.update(metadata)
    return {
        "type": block_type,
        "text": text,
        "page": page,
        "resource_ref": None,
        "source_ref": source_ref,
        "data": data,
    }


def _base_result(
    file_id: int,
    ext: str,
    blocks: list[dict[str, object]],
    warnings: list[str],
    metadata: dict[str, Any],
) -> dict[str, object]:
    return {
        "schema_version": SCHEMA_VERSION,
        "content_type": "document",
        "title": f"{ext.upper()} spreadsheet",
        "source_file_id": file_id if file_id > 0 else None,
        "source_module": SOURCE_MODULE,
        "parser": PARSER_NAME,
        "source": _source(file_id, ext),
        "file_id": file_id,
        "format": ext,
        "blocks": blocks,
        "resources": [],
        "warnings": warnings,
        "metadata": {
            "parser": PARSER_NAME,
            "format": ext,
            **metadata,
        },
        "quality": {},
    }


def _empty_result(file_id: int, ext: str, warnings: list[str], metadata: dict) -> dict:
    source_ref = {
        "format": ext,
        "line_start": None,
        "line_end": None,
        "row_count": 0,
        "kind": "empty_file",
    }
    block = _block("paragraph", f"Empty {ext.upper()} file", source_ref)
    return _base_result(file_id, ext, [block], warnings, metadata)


def parse_csv_file(file_id: int, path: Path, ext: str = "csv") -> dict:
    raw = path.read_bytes()
    content = _decode_text_bytes(raw)
    reader = csv.reader(io.StringIO(content))
    rows: list[str] = []
    warnings: list[str] = []
    truncated = False
    first_line: int | None = None
    last_line: int | None = None
    max_columns = 0

    for row_index, row in enumerate(reader, start=1):
        row_text = " | ".join(cell.strip() for cell in row).rstrip()
        if row_text.strip():
            rows.append(row_text)
            if first_line is None:
                first_line = row_index
            last_line = row_index
            max_columns = max(max_columns, len(row))
        if len(rows) >= MAX_ROWS_PER_SHEET:
            truncated = True
            rows.append(f"[... truncated at {MAX_ROWS_PER_SHEET} rows]")
            break

    metadata = {
        "row_limit": MAX_ROWS_PER_SHEET,
        "rows_emitted": min(len(rows), MAX_ROWS_PER_SHEET),
        "truncated": truncated,
    }
    if not rows:
        warnings.append("empty_csv")
        return _empty_result(file_id, ext, warnings, metadata)
    if truncated:
        warnings.append("row_limit_reached")

    source_ref = {
        "format": ext,
        "line_start": first_line,
        "line_end": last_line,
        "row_start": first_line,
        "row_end": last_line,
        "range": _range_for_rows(first_line, last_line, max_columns),
    }
    blocks = [_block("table", "\n".join(rows), source_ref)]
    return _base_result(file_id, ext, blocks, warnings, metadata)


def parse_xlsx_file(file_id: int, path: Path) -> dict:
    warnings: list[str] = []
    blocks: list[dict[str, object]] = []
    sheet_metadata: list[dict] = []
    data_block_count = 0
    value_workbook = None
    formula_workbook = None

    try:
        value_workbook = load_workbook(path, read_only=True, data_only=True)
        formula_workbook = load_workbook(path, read_only=True, data_only=False)
        for sheet_index, sheet_name in enumerate(value_workbook.sheetnames, start=1):
            value_sheet = value_workbook[sheet_name]
            formula_sheet = formula_workbook[sheet_name]
            rows: list[str] = []
            truncated = False
            first_row: int | None = None
            last_row: int | None = None
            max_columns = 0

            paired_rows = zip_longest(
                value_sheet.iter_rows(values_only=True),
                formula_sheet.iter_rows(values_only=True),
                fillvalue=(),
            )
            for row_index, (value_row, formula_row) in enumerate(paired_rows, start=1):
                row_text = _row_to_text(value_row or (), formula_row or ())
                if row_text.strip():
                    rows.append(row_text)
                    if first_row is None:
                        first_row = row_index
                    last_row = row_index
                    max_columns = max(max_columns, len(value_row or ()), len(formula_row or ()))
                if len(rows) >= MAX_ROWS_PER_SHEET:
                    truncated = True
                    rows.append(f"[... truncated at {MAX_ROWS_PER_SHEET} rows]")
                    break

            source_ref = {
                "format": "xlsx",
                "sheet": sheet_name,
                "sheet_index": sheet_index,
                "page": sheet_index,
                "row_start": first_row,
                "row_end": last_row,
                "range": _range_for_rows(first_row, last_row, max_columns),
            }
            sheet_metadata.append({
                "name": sheet_name,
                "page": sheet_index,
                "rows_emitted": min(len(rows), MAX_ROWS_PER_SHEET),
                "truncated": truncated,
                "source_ref": source_ref,
            })
            if truncated:
                warnings.append(f"row_limit_reached:{sheet_name}")
            if rows:
                block_text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
                blocks.append(_block("table", block_text, source_ref, page=sheet_index))
                data_block_count += 1
            else:
                warnings.append(f"empty_sheet:{sheet_name}")
                blocks.append(
                    _block(
                        "paragraph",
                        f"[Sheet: {sheet_name}]\nEmpty sheet",
                        source_ref,
                        page=sheet_index,
                        metadata={"empty": True},
                    )
                )
    except (BadZipFile, InvalidFileException, OSError, KeyError, ValueError) as exc:
        raise SpreadsheetParseError(f"Failed to parse XLSX file: {exc}") from exc
    finally:
        if value_workbook is not None:
            value_workbook.close()
        if formula_workbook is not None:
            formula_workbook.close()

    if data_block_count == 0:
        warnings.append("empty_workbook")

    if not blocks:
        source_ref = {
            "format": "xlsx",
            "sheet": None,
            "sheet_index": None,
            "page": None,
            "row_start": None,
            "row_end": None,
            "range": None,
        }
        blocks.append(_block("paragraph", "Empty XLSX workbook", source_ref))

    return _base_result(
        file_id,
        "xlsx",
        blocks,
        warnings,
        {
            "sheet_count": len(sheet_metadata),
            "sheets": sheet_metadata,
            "row_limit": MAX_ROWS_PER_SHEET,
        },
    )


def parse_spreadsheet_file(file_id: int, path: Path, ext: str) -> dict:
    normalized_ext = ext.lower().lstrip(".")
    if normalized_ext == "xlsx":
        return parse_xlsx_file(file_id, path)
    if normalized_ext == "csv":
        return parse_csv_file(file_id, path, normalized_ext)
    raise SpreadsheetParseError(
        f"Unsupported format '{normalized_ext}'. Allowed: {', '.join(sorted(SUPPORTED_EXTS))}"
    )


def _range_for_rows(first_row: int | None, last_row: int | None, max_columns: int) -> str | None:
    if first_row is None or last_row is None:
        return None
    end_column = get_column_letter(max(max_columns, 1))
    return f"A{first_row}:{end_column}{last_row}"
