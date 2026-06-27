"""FastAPI router for xlsx-parser module.

Registers the parse capability with the framework's cross-module registry.
"""
import os
from fastapi import APIRouter, Depends
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import AsyncSessionLocal, get_db
from app.middleware.auth import require_permission
from app.models.user import User
from app.schemas.common import ApiResponse
from app.services.module_registry import register_capability

router = APIRouter(prefix="/api/xlsx-parser", tags=["xlsx-parser"])


class ParseRequest(BaseModel):
    file_id: int


def _resolve_user_id(caller: str) -> int:
    from app.core.exceptions import PermissionDenied

    try:
        prefix, raw_id = caller.split(":", 1)
        if prefix == "user":
            return int(raw_id)
    except (TypeError, ValueError):
        pass
    raise PermissionDenied("Invalid caller")


async def _parse(params: dict, caller: str) -> dict:
    """Parse XLSX/CSV file into unified content blocks."""
    file_id = int(params.get("file_id", 0))
    if file_id <= 0:
        raise ValueError("file_id must be a positive integer")

    from app.config import get_settings
    from app.core.exceptions import NotFound, ValidationError, AppException
    from app.services.file_service import check_file_access
    from pathlib import Path
    import csv as csv_module
    import io
    import openpyxl

    allowed = {"xlsx", "xls", "csv"}
    user_id = _resolve_user_id(caller)
    async with AsyncSessionLocal() as db:
        file = await check_file_access(db, file_id, user_id)
        ext = (file.extension or "").lower()
        if ext not in allowed:
            raise ValidationError(f"Unsupported format '{ext}'. Allowed: {', '.join(sorted(allowed))}")
        if not file.storage_path:
            raise NotFound("File storage path is empty")
        upload_root = Path(get_settings().UPLOAD_DIR).resolve()
        full_path = (upload_root / file.storage_path).resolve()
        if os.path.commonpath([str(upload_root), str(full_path)]) != str(upload_root):
            raise AppException("Unsafe file storage path", status_code=400)
        if not full_path.exists() or not full_path.is_file():
            raise NotFound("File on disk not found")

        blocks = []

        if ext in ("xlsx", "xls"):
            wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    row_text = " | ".join(cells)
                    if row_text.strip():
                        rows.append(row_text)
                        row_count += 1
                    if row_count > 5000:
                        rows.append("[... truncated at 5000 rows]")
                        break
                if rows:
                    block_text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
                    blocks.append({"type": "表格", "text": block_text, "page": None, "resource_ref": None})
            wb.close()
        elif ext == "csv":
            raw = full_path.read_bytes()
            try:
                content = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                content = raw.decode("gbk", errors="replace")
            reader = csv_module.reader(io.StringIO(content))
            rows = []
            for i, row in enumerate(reader):
                cells = [c.strip() for c in row]
                rows.append(" | ".join(cells))
                if i >= 5000:
                    rows.append("[... truncated at 5000 rows]")
                    break
            if rows:
                blocks.append({"type": "表格", "text": "\n".join(rows), "page": None, "resource_ref": None})

    return {
        "file_id": file_id,
        "format": ext,
        "blocks": blocks,
        "resources": [],
    }


@router.get("/health")
async def health():
    return ApiResponse(data={"module": "xlsx-parser", "status": "ok"})


@router.post("/parse")
async def call_parse(payload: ParseRequest, user: User = Depends(require_permission("viewer"))):
    result = await _parse({"file_id": payload.file_id}, f"user:{user.id}")
    return ApiResponse(data=result)


register_capability(
    "xlsx-parser", "parse", _parse,
    description="Parse XLSX/CSV files into unified content blocks",
    brief="解析 XLSX 文档",
    parameters={"file_id": {"type": "int"}},
    min_role="viewer",
)
