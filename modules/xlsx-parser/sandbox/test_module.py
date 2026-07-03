"""Sandbox test for xlsx-parser module."""
from __future__ import annotations

import importlib.util
import tempfile
from pathlib import Path

from openpyxl import Workbook

SDIR = Path(__file__).resolve().parent / "samples"
CORE_PATH = Path(__file__).resolve().parents[1] / "backend" / "parser_core.py"
CORE_SPEC = importlib.util.spec_from_file_location("xlsx_parser_core", CORE_PATH)
if CORE_SPEC is None or CORE_SPEC.loader is None:
    raise RuntimeError("Failed to load parser core")
parser_core = importlib.util.module_from_spec(CORE_SPEC)
CORE_SPEC.loader.exec_module(parser_core)

MAX_ROWS_PER_SHEET = parser_core.MAX_ROWS_PER_SHEET
SpreadsheetParseError = parser_core.SpreadsheetParseError
parse_spreadsheet_file = parser_core.parse_spreadsheet_file


def validate(result: dict[str, object], label: str) -> None:
    assert all(k in result for k in ("file_id", "format", "blocks", "resources"))
    assert isinstance(result["blocks"], list)
    for block in result["blocks"]:
        assert isinstance(block, dict)
        assert all(k in block for k in ("type", "text", "page", "resource_ref"))
    print("  [%s] Validation PASS (%d blocks)" % (label, len(result["blocks"])))


def _parse_sample(filename: str) -> dict[str, object]:
    sample = SDIR / filename
    assert sample.exists(), f"Sample not found: {sample}"
    return parse_spreadsheet_file(0, sample, sample.suffix.lstrip("."))


def _write_formula_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Formula"
    sheet.append(["A", "B", "Total"])
    sheet.append([10, 20, "=A2+B2"])
    workbook.save(path)


def _write_empty_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "Empty"
    workbook.save(path)


def _write_large_workbook(path: Path) -> None:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Large")
    for index in range(MAX_ROWS_PER_SHEET + 2):
        sheet.append([index, f"row-{index}"])
    workbook.save(path)


def _assert_formula_boundary() -> None:
    with tempfile.TemporaryDirectory(prefix="xlsx-parser-formula-") as temp_dir:
        path = Path(temp_dir) / "formula.xlsx"
        _write_formula_workbook(path)
        result = parse_spreadsheet_file(0, path, "xlsx")
    text = "\n".join(block["text"] for block in result["blocks"])
    assert "=A2+B2" in text, "Formula text should be preserved when no cached value exists"
    validate(result, "formula.xlsx")


def _assert_empty_boundary() -> None:
    with tempfile.TemporaryDirectory(prefix="xlsx-parser-empty-") as temp_dir:
        path = Path(temp_dir) / "empty.xlsx"
        _write_empty_workbook(path)
        result = parse_spreadsheet_file(0, path, "xlsx")
    assert result["blocks"] == []
    assert "empty_workbook" in result.get("warnings", [])
    validate(result, "empty.xlsx")


def _assert_large_boundary() -> None:
    with tempfile.TemporaryDirectory(prefix="xlsx-parser-large-") as temp_dir:
        path = Path(temp_dir) / "large.xlsx"
        _write_large_workbook(path)
        result = parse_spreadsheet_file(0, path, "xlsx")
    text = "\n".join(block["text"] for block in result["blocks"])
    assert f"[... truncated at {MAX_ROWS_PER_SHEET} rows]" in text
    assert any(str(item).startswith("row_limit_reached") for item in result.get("warnings", []))
    validate(result, "large.xlsx")


def _assert_bad_file_fails() -> None:
    with tempfile.TemporaryDirectory(prefix="xlsx-parser-bad-") as temp_dir:
        path = Path(temp_dir) / "bad.xlsx"
        path.write_text("not an xlsx file", encoding="utf-8")
        try:
            parse_spreadsheet_file(0, path, "xlsx")
        except SpreadsheetParseError:
            print("  [bad.xlsx] Failure PASS")
        else:
            raise AssertionError("Bad XLSX should raise SpreadsheetParseError")


def main() -> None:
    print("=" * 60)
    print("xlsx-parser sandbox test")
    print("=" * 60)
    for fn in ("sample.xlsx", "sample.csv"):
        result = _parse_sample(fn)
        for block in result["blocks"]:
            print("  [%s] %s" % (fn, block["text"][:80]))
        validate(result, fn)

    sample_xlsx = _parse_sample("sample.xlsx")
    assert len(sample_xlsx["blocks"]) >= 2, "Expected multi-sheet XLSX sample"
    assert sample_xlsx["blocks"][0]["page"] == 1
    assert sample_xlsx["blocks"][1]["page"] == 2

    _assert_formula_boundary()
    _assert_empty_boundary()
    _assert_large_boundary()
    _assert_bad_file_fails()
    print("PASS: xlsx-parser sandbox test")


def test_xlsx_parser_sandbox_contract() -> None:
    main()


if __name__ == "__main__":
    main()
