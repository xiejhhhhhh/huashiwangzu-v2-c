import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from app.core.exceptions import ValidationError

logger = logging.getLogger(__name__)


class ExcelService:

    MAX_FILE_SIZE = 10 * 1024 * 1024

    async def parse(self, file_path: str, file_size: int) -> dict:
        if file_size > self.MAX_FILE_SIZE:
            mb = round(file_size / 1024 / 1024, 1)
            raise ValidationError(f"文件过大（{mb}MB），仅支持 10MB 以内的文件")

        wb = load_workbook(file_path, data_only=True)
        sheets = []
        for ws in wb.worksheets:
            cells = {}
            formulas = []
            for row in ws.iter_rows():
                for cell in row:
                    coord = f"{get_column_letter(cell.column)}{cell.row}"
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas.append({
                            "cell": coord, "formula": cell.value,
                            "cached_result": str(cell.value),
                        })
                        cells[coord] = {"value": str(cell.value), "type": "formula"}
                    elif cell.value is None:
                        cells[coord] = {"value": None, "type": "null"}
                    elif isinstance(cell.value, (int, float)):
                        cells[coord] = {"value": cell.value, "type": "number"}
                    else:
                        cells[coord] = {"value": str(cell.value), "type": "string"}

            sheets.append({
                "name": ws.title,
                "index": wb.worksheets.index(ws),
                "rows": ws.max_row,
                "cols": ws.max_column,
                "cells": cells,
                "formulas": formulas,
            })

        return {
            "version": "1.0",
            "metadata": {
                "title": wb.properties.title or "",
                "sheet_count": len(wb.sheetnames),
                "active_sheet": 0,
            },
            "sheets": sheets,
        }

    async def export(self, file_path: str, json_data: dict) -> None:
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)

        sheets = json_data.get("sheets", [])
        if not sheets:
            inner = json_data.get("content", {})
            sheets = inner.get("sheets", []) if inner else []

        for sheet_data in sheets:
            ws = wb.create_sheet(title=sheet_data.get("name", "Sheet1"))
            cells = sheet_data.get("cells", {})
            for coord, info in cells.items():
                if info.get("type") == "null":
                    continue
                ws[coord] = info.get("value")

        wb.save(file_path)
